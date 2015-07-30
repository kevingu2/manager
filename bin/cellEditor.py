import sys
from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string
from editpyxl import Workbook
def findRows(fileName, ids):
    wb = load_workbook(filename = fileName, read_only=True, keep_vba=True, use_iterators=True)
    ws = wb.get_sheet_by_name(name='PipelineView')
    count = 1
    rows = []
    for i in ids:
        count = 1
        for row in ws.iter_rows():
            #print count
            if row[0].value == i:
                rows.append(count)
                break
            else:
                count += 1
                continue
    return rows

def findCols(cols):
    # can only edit certain columns. if more are needed, add to this dictionary
    column_hash = {'slDir': 'CZ', 'slArch': 'DA', 'leadEstim': 'DB', 'engaged': 'DL', 'solution': 'DM', 'estimate': 'DN', 'slComments': 'DO'}
    columns = []
    for c in cols:
        columns.append(column_index_from_string(column_hash[c]))
    return columns


print "*"*10, "starting python excel editor", "*"*10
args = sys.argv
fileName = args[1]
strs = args[2]
wb = load_workbook(filename = fileName, keep_vba = True)
ws = wb.get_sheet_by_name('PipelineView')
data = [[i for i in x.strip(" []").split(", ")] for x in strs.strip('[]').split("],")] # magic
#####get the rows/cols/valuesToChange#########
ids = []
cols = []
changes = []
columns = []
print "*"*10, "parsing data", "*"*10
for d in data:
    ids.append(d[0].strip('\''))
    cols.append(d[1].strip('\''))
    changes.append(d[2].strip('\''))

rows = findRows(fileName, ids)
cols = findCols(cols)
print "*"*10, "starting editing", "*"*10
for c in range(len(changes)):
    ws.cell(row=rows[c], column=cols[c]).value = changes[c]
print "*"*10, "finished editing", "*"*10
wb.save(fileName)
print "*"*10, "finished saving", "*"*10
