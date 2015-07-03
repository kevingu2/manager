from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
import win32com.client
from win32com.client import Dispatch
from oletools.olevba import VBA_Parser, VBA_Scanner
import glob
import random
import re
def extractVBA(fileName):
    codeList = []
    vba = VBA_Parser('BD-PipelineViewer8 13 April.xlsm')
    if not os.path.exists("scripts"):
        os.makedirs("scripts")
    for (filename, stream_path, vba_filename, vba_code) in vba.extract_macros():
        codeList.append(vba_code)

    return codeList

def edit(fileName, row, data):
    wb = load_workbook(fileName, keep_vba=True)
    sheet_names = wb.get_sheet_names() # list of all sheet names
    ws = wb.get_sheet_by_name(sheet_names[0]) # get first sheet
    ws['F1'] = "Some Value"
    wb.save('b.xlsm')
    codeList = extractVBA(fileName)

    _file = os.path.abspath(sys.argv[0])
    path = os.path.dirname(_file)
    pathToExcelFile = path + '\\b.xlsm'
    pathToMacro = path + '\scripts\Module32.bas'
    MyMacroName = 'FuckThis'
    #pathToExcelFile.st

    # read the textfile holding the excel macro into a string


    # open up an instance of Excel with the win32com driver
    excel = win32com.client.Dispatch("Excel.Application")

    # do the operation in background without actually opening Excel
    excel.Visible = False

    # open the excel workbook from the specified file
    #workbook = excel.Workbooks.Open(Filename=pathToExcelFile, CorruptLoad=1)
    workbook = win32com.client.GetObject(pathToExcelFile)

    for macro in codeList:
        # insert the macro-string into the excel file
        excelModule = workbook.VBProject.VBComponents.Add(1)
        excelModule.CodeModule.AddFromString(macro)

    # save the workbook and close
    excel.Workbooks(1).Close(SaveChanges=1)
    excel.Application.Quit()
#below is an example of the usage
#edit('abc.xlsm', 3,"{a:123, b:234, c:345}")
#      fileName, row, newData

args = sys.argv # gets args from command line
dict = {'A':1, 'B':2, 'C':3, 'D':4, 'E':5, 'F':6}
edit(args[1], int(args[2]), dict)

#view = View(args[1])
#print_data(view[0])
