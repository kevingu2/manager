import xlrd
import sys
from editpyxl import Workbook
import ast
from openpyxl import load_workbook
import json


args = sys.argv # get command line args
wb = Workbook()
fileName = args[1]
wb.open(fileName)
ws = wb.get_sheet_by_name('Admin')

print ws.cell("M2").value
